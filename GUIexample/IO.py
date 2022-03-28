import time
from datetime import datetime
import threading
import numpy as np
from NI_DAQ import AIport, AOport, AOport2
from misc import calc_temp
from spectroscopy import Spectroscopy
from langmuir import Probe
from breakdown import Breakdown
from misc import console_out, tick, tock, Benchmark
from PyDAQmx import DAQError
from LJ_DAQ import LJMport
from labjack import ljm
import visa
import pickle
import openpyxl
import string
from simple_pid import PID
import pidcontroller

class IO:
    
    def __init__(self, path):
        self.path = path
        self.shutdown = None
        self.IO_OK = False
        self.log_T0 = time.time()
        self.I_max = 10
        self.P_max = 15
        self.OC = False
        self.RM = visa.ResourceManager()
        self.FIL_VOL = 0
        self.ARC_VOL = 0
        self.openvalve_pressurebase = 25

        
        console_out("Opening GPIB-USB port")
        try:
            self.FIL_PS = self.RM.open_resource('COM4')
            self.FIL_PS.write('FOLD CV')
            self.FIL_PS_OK = True
        except:     
            console_out("GPIB-USB could not be found", error=True)
            self.FIL_PS_OK = False
        
        console_out("Opening NI_DAQ ports")
        try:
            self.AI_ARC_VOL_A_NI = AIport('/Dev3/ai0')
            self.AI_ARC_VOL_B_NI = AIport('/Dev3/ai1')
            self.NI_DAQ_OK = True
        except DAQError:
            console_out("NI DAQ could not be found", error=True)
            self.NI_DAQ_OK = False

        console_out("Opening LJM ports")
        self.AO_ARC = LJMport('DAC1')       #Inlet Arc Voltage
        self.AO_FIL = LJMport('DAC0')       #Inlet Fil Current  #Not anymore, change with read function of 
        
        self.AI_LANG_VOL = LJMport('AIN1', _res=5)
        self.AI_LANG_CUR = LJMport('AIN3', _range=10, _res=4)
        self.AI_LANG_CUR_B = LJMport('AIN2', _range=10, _res=4)
        self.AI_LANG_CUR_A = LJMport('AIN3', _range=10, _res=4)

        self.AI_BD_VOL = LJMport('AIN0', _res=5)
        self.AI_BD_CUR = LJMport('AIN2', _range=10, _res=4)
        
        self.AI_P1 = LJMport('AIN48')
        
        console_out("Reading thermocouples")
        self.AI_T_A1 = LJMport('AIN50', _range=10, _res=4)      # Channel A, type R
        self.AI_T_A2 = LJMport('AIN51', _range=10, _res=4)      # Channel B, type E   

        self.AI_ARC_VOL_A = LJMport('AIN56')
        self.FACTOR_A = 88.87
        self.AI_ARC_VOL_B = LJMport('AIN57')
        self.FACTOR_B = 90.15
        
        # Shunt resistor of 10 Ohms and amplifier at gain of 100
        self.AI_LANG2_CURR = LJMport('AIN58', _range=10, _res=4)    

        self.AI_ARC_CUR = LJMport('AIN55')
        self.AI_FIL_CUR = LJMport('AIN54')      #Measured
        
        console_out("Opening spectrometers")
        try:
            self.spectroscopy = Spectroscopy()
            self.spectra_timestamps = []
            self.SPEC_OK = True
        except ValueError as err:
            self.SPEC_OK = False
            console_out(err.args[0], error=True)

        self.probe = Probe()
        self.breakdown = Breakdown()
                
        self.acquisition_instance = threading.Thread(target=self.acquisition_thread, daemon=True)        
        self.acquisition_instance.logging = False
        if(self.SPEC_OK):
            self.spectroscopy_acquisition_instance = threading.Thread(target=self.spectroscopy_acquisition_thread, daemon=True)
            self.spectroscopy_acquisition_instance.logging = False
        if(self.NI_DAQ_OK):
            self.voltage_acquisition_instance = threading.Thread(target=self.voltage_acquisition_thread, daemon=True)
            
        self.names = [self.AI_LANG_VOL.name, self.AI_BD_CUR.name, self.AI_LANG_CUR.name, self.AI_P1.name, self.AI_ARC_VOL_A.name, self.AI_ARC_VOL_B.name, self.AI_ARC_CUR.name, self.AI_FIL_CUR.name, self.AI_T_A1.name, self.AI_T_A2.name, self.AI_LANG2_CURR.name, self.AI_BD_VOL.name]
        #self.names = [self.AI_LANG_VOL.name, self.AI_LANG_CUR_A.name, self.AI_LANG_CUR_B.name, self.AI_P1.name, self.AI_ARC_VOL_A.name, self.AI_ARC_VOL_B.name, self.AI_ARC_CUR.name, self.AI_FIL_CUR.name, self.AI_T_A1.name, self.AI_T_A2.name, self.AI_LANG2_CURR.name]
        self.device = ljm.openS()


        #SIM: STREAM IS RUNNING ERROR FIX (APPEARS WHEN PROGRAM IS NOT CLOSED SUFFIECIENT
        try:
           LJMport.get('AIN1')
        except ljm.LJMError:
           ljm.eStreamStop(self.device) #THIS LINE FOR ERROR 2605 WITH NO INDENT
        #END ERROR FIX


                    
        LJMport.set('STREAM_BUFFER_SIZE_BYTES', 32768)
        
        self.stream_length = 50
        self.stream_rate = 20

        LJMport.set('STREAM_RESOLUTION_INDEX', 0)
        LJMport.set('STREAM_SETTLING_US', 0)
        ljm.eStreamStart(self.device, self.stream_length, len(self.names), [ljm.nameToAddress(name)[0] for name in self.names], self.stream_length*self.stream_rate)

        self.spec_aq_bench = Benchmark()
        self.aq_bench = Benchmark()
        self.langmuir_aq_bench = Benchmark()


    def reset(self):
        self.AO_ARC.write(0)
        self.FIL_PS.write('CLR')
        self.AO_FIL.write(0)

    def start(self):
        self.acquisition_instance.start()
        if(self.SPEC_OK):
            self.spectroscopy_acquisition_instance.start()
        if(self.NI_DAQ_OK):
            self.voltage_acquisition_instance.start()
    
    def spectroscopy_acquisition_thread(self):
        current_thread = threading.currentThread()

        self.spectra_log1 = np.array([]).reshape(0,2)
        self.spectra_log2 = np.array([]).reshape(0,2)
        self.spectra_timestamps = []
        
        while(getattr(current_thread, "active", True)):
            self.spec_aq_bench.mark()
            if(self.SPEC_OK):
                self.spectra = np.array(self.spectroscopy.acquire_all())

                if(getattr(current_thread, "logging", True)):
                    self.spectra_timestamps = np.append(self.spectra_timestamps, self.T-self.log_T0)

                    self.spectra_log1 = np.append(self.spectra_log1, self.spectra[0], axis=0)

                    if(len(self.spectra) > 1):
                        self.spectra_log2 = np.append(self.spectra_log2, self.spectra[1], axis=0)
                                        
    def start_probe(self):
        self.probe.saving = True
        self.breakdown.saving = True
        
    def stop_probe(self):
        self.probe.saving = False
        self.breakdown.saving = False
        self.probe.save(self.shotname)
        self.breakdown.save(self.shotname)
        
    
    def save_spectra(self, dark=False, back=False):
        if(not self.SPEC_OK):
            return
        
        filename = 'data\\' + self.shotname + '\\SPECT1.p'        
        if(dark):
            filename = 'data\\' + self.shotname + '\\SPECT1_DARK.p'
        if(back):
            filename = 'data\\' + self.shotname + '\\SPECT1_BACK.p'
        d1 = {'WAVELENGTH': self.spectra[0][:,0], 'INTENSITY': self.spectra[0][:,1]}          
        pickle.dump(d1, open(filename, 'wb'))

        
        if(len(self.spectra) > 1):
            
            filename = 'data\\' + self.shotname + '\\SPECT2.p'
            if(dark):
                filename = 'data\\' + self.shotname + '\\SPECT2_DARK.p'
            if(back):
                filename = 'data\\' + self.shotname + '\\SPECT2_BACK.p'
            d2 = {'WAVELENGTH': self.spectra[1][:,0], 'INTENSITY': self.spectra[1][:,1]}
            pickle.dump(d2, open(filename, 'wb'))

    def save_spectra_log(self):
        filename = 'data\\' + self.shotname + '\\SPECT1LOG.p'  
        d1 = {'WAVELENGTH': self.spectra_log1[:,0], 'INTENSITY': self.spectra_log1[:,1], 'TIMESTAMPS': self.spectra_timestamps}          
        pickle.dump(d1, open(filename, 'wb'))
        
        if(len(self.spectra) > 1):
            filename = 'data\\' + self.shotname + '\\SPECT2LOG.p'  
            d2 = {'WAVELENGTH': self.spectra_log2[:,0], 'INTENSITY': self.spectra_log2[:,1], 'TIMESTAMPS': self.spectra_timestamps}          
            pickle.dump(d2, open(filename, 'wb'))

    def save_time_points(self, t0,tf):
        ### Save set points in a .p file
        filename = 'data\\' + self.shotname + '\\TIMEPOINTS.p'
        d2 = {'T0': t0, 'TF': tf}
        pickle.dump(d2, open(filename, 'wb'))
        
    def save_setpoints(self, text, dat):
        ### Save set points in a .p file
        filename = 'data\\' + self.shotname + '\\SETPOINTS.p'
        d3 = {'DATETIME': datetime.now().strftime('%m-%d-%Y_%H%M'), 'SHOTNUMBER': self.shotname, 'SHOTTYPE': text[0], 'FILLCURR': text[3], 'ARCVOL': text[2], 'AV0': dat[0], 'AVF': dat[1], 'FC0': dat[2], 'FCF': dat[3], 'PROBC': dat[4], 'TIME': dat[5], 'PERIOD': dat[6], 'STABTIME': dat[7], 'COMMENTS' : text[1], 'PRESTYPE': text[4], 'PRESINT': dat[8], 'PRESFIN': dat[9]}
        pickle.dump(d3, open(filename, 'wb'))

        ### Save set points in a .xlsx file
        try:
            wbook = openpyxl.load_workbook('Input Data.xlsx')
            wsheet = wbook['Sheet1']
            row = str(int(self.shotname)-14)
            col = list(string.ascii_uppercase)
            i = 0
            for key in d3:
                              
                if key == 'STABTIME':
                    pass
                else:
                    if d3[key] == 0 :
                        d3[key] = "--"
                    cell = col[i] + row
                    wsheet[cell] = d3[key]
                    i += 1 
            wbook.save('Input Data.xlsx')

        except:
            import os
            os.system("taskkill /im excel.exe")
            print('EXCEL FILES OPEN, please save the files within 60 seconds else the alternations will be lost')
            print('Please wait...')
            time.sleep(60)
            os.system("taskkill /f /im excel.exe")
            wbook = openpyxl.load_workbook('Input Data.xlsx')
            wsheet = wbook['Sheet1']
            row = str(int(self.shotname)-14)
            col = list(string.ascii_uppercase)
            i = 0
            for key in d3:
                              
                if key == 'STABTIME':
                    pass
                else:
                    if d3[key] == 0 :
                        d3[key] = "--"
                    cell = col[i] + row
                    wsheet[cell] = d3[key]
                    i += 1 
            wbook.save('Input Data.xlsx')
            
    def voltage_acquisition_thread(self):
        current_thread = threading.currentThread()
        
        while(getattr(current_thread, "active", True)):
            self.langmuir_aq_bench.mark()
            
            VOLA = (self.AI_ARC_VOL_A_NI.read()+0.59)*75/0.47
            VOLB = (self.AI_ARC_VOL_B_NI.read()+0.59)*75/0.47
            
            self.ARC_VOL = VOLB
            self.FIL_VOL = np.abs(VOLA-VOLB)
            
    def acquisition_thread(self):
        current_thread = threading.currentThread()
        
        while(getattr(current_thread, "active", True)):
            self.aq_bench.mark()
            
            self.T = time.time()

            self.T1 = 0

            def read_data():
                try:
                    return np.transpose(np.array(ljm.eStreamRead(self.device)[0]).reshape(self.stream_length, len(self.names)))
                except:
                    console_out("Buffer full, restarting stream")
                    ljm.eStreamStop(self.device)
                    ljm.eStreamStart(self.device, self.stream_length, len(self.names), [ljm.nameToAddress(name)[0] for name in self.names], self.stream_length*self.stream_rate)        
                    return np.transpose(np.array(ljm.eStreamRead(self.device)[0]).reshape(self.stream_length, len(self.names)))
                
            dat = read_data()


            ### Pressure conversion to Torr
            ###############################
            presstemp = (10**(1.667*(dat[3,-1])-11.46))
##BRUTE FORCING IT
            #if (presstemp > 0.025):
                #self.pressure = presstemp + 0.0015 + (presstemp-0.025)/10
            #elif (presstemp > 0.009):
                #self.pressure = presstemp + 0.0015 - (0.026-presstemp)/10
            #else:
                #self.pressure = presstemp
##DATA FROM EXCEL
            if (presstemp > 0.008):
                self.pressure = 1.0881*presstemp - 0.0002
            else:
                self.pressure = presstemp
            
            #correctionpar = 0.0032*(self.pressure-0.025)**2
            #correctionhyp = (np.exp(0.05*(self.pressure*100-25)-1))/100
            #self.pressure = 0.001*pressint+0.0015
            #self.pressure = correctionfit
            #if (self.pressure > 0.02) and (self.pressure < 0.035):
            #    self.pressure = self.pressure+0.002
            #elif (self.pressure>0.035) and (self.pressure<0.055):
            #    self.pressure = self.pressure + 0.004
            #elif (self.pressure>0.065) and (self.pressure<0.055):
            #    self.pressure = self.pressure - correctionhyp                       

            #GONZALO
            #CHECK THE 0.004 AT THE END BECAUSE WE ADDED THAT
            #print(self.pressure)


            ### Filament and arc setp and input/measured values
            self.FIL_CUR_SETP = self.AO_FIL.value*(120/5)     # CHECK IF THIS CHANGE IS CORRECT       
            self.ARC_VOL_SETP = self.AO_ARC.value*(160/5)
            
            self.ARC_CUR = -dat[6,-1]/500*1000
            self.FIL_CUR = dat[7,-1]/75*1000


            ### Thermocouples conversion to Celsius
            ###############################
            #self.mvolt1
            self.ThTR1 = (dat[8,-1] - 0.399)/201
            #self.mvolt2
            self.ThTE2 = (dat[9,-1] - 0.4)/51
            # Thermocouple Type Constants
            #ttB = 6001  #ttE = 6002  #ttJ = 6003
            #ttK = 6004  #ttN = 6005  #ttR = 6006
            #ttS = 6007  #ttT = 6008  #ttC = 6009
            #Notes:
            #B-type measurements below ~373 degrees Kelvin or ~0.04
            #millivolts (at a cold junction junction temperature of 273.15
            #degrees Kelvin) may be inaccurate.

            # Temperature of LJTIA is 22.4 C -> 295.55 K
            
            #self.ThTR1 = ljm.tcVoltsToTemp(6006, self.mvolt1, 295) - 273.15 - 14 #Offset
            #self.ThTE2 = ljm.tcVoltsToTemp(6002, self.mvolt2, 295) - 273.15


            ### Resistor shunt for measuring current on probe biased negative
            ###############################
            self.currshunt = dat[10,-1]     #mA
            #self.currshunt = dat[8,-1]     #mA
    

            ### Langmuir probe traces
            ###############################
            
            #self.probe.update_array(10*dat[0,:], 0.4*dat[1,:]/100, 0.4*dat[2,:]/10, self.log_T0)
            #self.probe.update_array(10*dat[0,:], 0.1*dat[1,:], 0.1*dat[2,:]/10, self.log_T0) 

            R = 100000           # Resistance on output (Ohm)
            Ioff = -.000030           # Offset current
            #IoffB = 0.000490
            probe_vol = 10.86071*dat[0,:]
            #probe_currA = 0.1*dat[1,:]/1 - 0.001*probe_vol/R + IoffA        # Amplified 100 times     #Purple
            #probe_curr = 0.1*dat[2,:]- probe_vol*1.05e-5 + 645*1e-6
            probe_curr = 0.1*dat[2,:]- probe_vol*1.05e-5 +150*1e-6#+ 645*1e-6

            # No Amplified
            self.probe.update_array(probe_vol, probe_curr, probe_curr, self.log_T0)


            ## Electrical Breakdown test
            ###############################
            R = 100000           # Resistance on output (Ohm)
            Voff = 0.185
            Ioff = 0.0096-0.0068
            #self.bd_vol = 34.7971*(dat[9,:] + Voff)
            bd_vol = 34.7971*(dat[11,:] + Voff)
            #probe_currA = 0.4*dat[1,:] - probe_vol/R + Ioff        # Amplified 100 times     #Purple
            bd_curr = 0.4*dat[1,:] + Ioff # - self.bd_vol/R        #    #Green
            #self.probe.update_array(self.bd_vol, self.bd_curr, self.bd_curr, self.log_T0)
            self.breakdown.update_array(bd_vol, bd_curr, bd_curr, self.log_T0)

            
            if(self.ARC_CUR > self.I_max):
                self.reset()
                self.OC = True
                console_out("ERROR-> OVERCURRENT AT %.1f A" % self.I_max)
                 
            def tungsten_resistivity_inverse(rho):
                return 2000+(rho*1e6*1e2-56.83)/0.03582 #K
            rho_fil = (np.pi * (1.5e-3 / 2)**2 / 160e-3) * (self.FIL_VOL/self.FIL_CUR)
            
            self.T_FIL = np.max([tungsten_resistivity_inverse(rho_fil), 0])
            
            if(getattr(current_thread, "logging", True)):
                self.data_log['TIME'].append(self.T-self.log_T0)
                self.data_log['PRESSURE'].append(self.pressure)
                self.data_log['FIL_CUR_SETP'].append(self.FIL_CUR_SETP)
                self.data_log['ARC_VOL_SETP'].append(self.ARC_VOL_SETP)
                self.data_log['FIL_CUR'].append(self.FIL_CUR)
                self.data_log['ARC_VOL'].append(self.ARC_VOL)
                self.data_log['ARC_CUR'].append(self.ARC_CUR)
                self.data_log['FIL_VOL'].append(self.FIL_VOL)
                self.data_log['CUR_SHUNT'].append(self.currshunt)
                #Provitional
                self.data_log['TEMPERATURE'].append(self.ThTR1)

            self.IO_OK = True
            """
            if self.aq_bench.FPS < 10:
                console_out("Buffer full, restarting stream")
                ljm.eStreamStop(self.device)
                ljm.eStreamStart(self.device, self.stream_length, len(self.names), [ljm.nameToAddress(name)[0] for name in self.names], self.stream_length*self.stream_rate)  
                pass
            """
   
    def start_log(self):
        self.timestamp = datetime.now()
        self.log_T0 = time.time()
        self.data_log = {'TIME': [], 'PRESSURE': [], 'FIL_CUR_SETP': [], 'FIL_CUR': [], 'FIL_VOL': [], 'ARC_VOL_SETP': [], 'ARC_VOL': [], 'ARC_CUR': [], 'CUR_SHUNT': [], 'TEMPERATURE' : []}
        self.acquisition_instance.logging = True
        #self.shotname =  datetime.now().strftime('%m-%d-%Y_%H%M')
        
        ### Reading folders' name to set the next shotname
        import os
        try: 
            for root, dirs, files in os.walk("data\\", topdown=False):
                for i in np.arange(0, len(dirs)):
                    row = int(dirs[i])
            self.shotname = "{:05d}".format(row+1)      # Takes the following shot number
        except:     #There are no files on folder
            self.shotname = "00001"
            ### Create a excel file to insert input data
            import xlsxwriter
            wbook = xlsxwriter.Workbook('Input Data.xlsx')
            wsheet = wbook.add_worksheet()
            
            ### Formatting excel file
            bold = wbook.add_format({'bold':True})
            cell_format = wbook.add_format()
            cell_format.set_align('right')
            wsheet.set_column('A:A', 15)
            wsheet.set_column('B:B', 14)
            wsheet.set_column('C:C', 17)
            wsheet.set_column('D:D', 17)
            wsheet.set_column('E:E', 17)
            wsheet.set_column('K:K', 12)
            wsheet.set_column('M:M', 75)

            ### Creating titles
            wsheet.write(0, 0, "DATE TIME", bold)
            wsheet.write(0, 1, "SHOT NUMBER", bold)
            wsheet.write(0, 2, "TYPE OF SHOT", bold)
            wsheet.write(0, 3, "FIL CURR", bold)
            wsheet.write(0, 4, "ARC VOL", bold)
            wsheet.write(0, 5, "Varc0 (V)", bold)
            wsheet.write(0, 6, "Varcf (V)", bold)
            wsheet.write(0, 7, "Ifil0 (I)", bold)
            wsheet.write(0, 8, "Ifilf (I)", bold)
            wsheet.write(0, 9, "Iprob (mA) ", bold)
            wsheet.write(0, 10, "DURATION(s)", bold)
            wsheet.write(0, 11, "PERIOD(s) ", bold)
            wsheet.write(0, 12, "COMMENTS", bold)
            wsheet.write(0, 13, "PRESS", bold)
            wsheet.write(0, 14, "GAS PRESSURE INITIAL (mTorr)", bold)
            wsheet.write(0, 15, "GAS PRESSURE FINAL (mTorr)", bold)
            
        try:
            os.stat('data\\' + self.shotname + '\\')
        except:
            os.mkdir('data\\' + self.shotname + '\\')
        self.spectroscopy_acquisition_instance.logging = True

    def del_logs(self):
        import shutil
        try:
            shutil.rmtree('data\\' + self.shotname + '\\')
        except:
            pass
                    
    def stop_log(self):
        self.acquisition_instance.logging = False
        self.spectroscopy_acquisition_instance.logging = False
        self.save_spectra_log()        
                    
        import pickle
        with open('data\\' + self.shotname + '\\SHOT.p', 'wb') as handle:
            pickle.dump(self.data_log, handle)
        
    def close(self):
        self.IO_OK = False
        
        self.acquisition_instance.active = False
        if(self.SPEC_OK):
            self.spectroscopy_acquisition_instance.active = False  
        if(self.NI_DAQ_OK):
            self.voltage_acquisition_instance.active = False
        #self.mass_check_thread.active=False
        time.sleep(1)

        #p.write(0)

        ljm.eStreamStop(self.device)
        
        ljm.closeAll()

    
    
    #SIM PRESSURE CONTROL (WORKING PERFECTLY :) )
    def Press_start(self,p,p_close,p0,pf):
        if (p0 != 0.0):
            mass_Flow = (p0/10)# - 0.2)
            pset=p0
        else:    
            mass_Flow = (pf/10)# - 0.2)
            pset=pf
        p_close.write(1)
        p.write(mass_Flow)
        while (abs(self.pressure*1000-pset)>0.5):
            time.sleep(1)
        return pset, mass_Flow
        
    def Press_control(self,p,pressswitch,presscontype,p0,pf,intmass,T):
        pid = pidcontroller.PID(3, 0, 0)
                
        while True: #(pressswitch==True):
            if pressswitch():
                console_out("Stop gas control")
                break
            pressfac=0.001
            mass_Flow=intmass
            mass_Step=0.05
            #console_out("Entered While")
            if (presscontype=='Stable'):
                pset=p0
                console_out("Starting Stabile gas control")
                while (presscontype=='Stable'):
                    
                    error = 0.001*pset - self.pressure
                    
                    correction = pid.Update(error)

                    mass_Flow = mass_Flow + correction
                    if (mass_Flow >4.99):
                            mass_Flow=5
                    if (mass_Flow < 0.01):
                            mass_Flow=0                    
                    p.write(mass_Flow)
                    
                    #print(mass_Flow)
                    intmass=mass_Flow
                    #console_out("Stable")
                    time.sleep(0.1)                    
                    if pressswitch():
                        console_out("Stop gas control")
                        break
                if pressswitch():
                    console_out("Stop gas control")
                    break
            if pressswitch():
                console_out("Stop gas control")
                break
            
                    
            elif (presscontype=='RAMP'):
                counts=500
                t=T
                pchange=pf+5-p0
                mass_Step=0.05
                console_out("Start Ramp Gas")
                for i in np.arange(0, counts + 1):

                    
                    pset=p0+i*(pchange)/counts

                    error = 0.001*pset - self.pressure
                    
                    correction = pid.Update(error)
                    

                    mass_Flow = mass_Flow + correction
                    if (mass_Flow >4.99):
                            mass_Flow=5
                    if (mass_Flow < 0.01):
                            mass_Flow=0                    
                    p.write(mass_Flow)
                    
                    #print(mass_Flow)
                    intmass=mass_Flow
                    #console_out("RAMP")
                   
                   
                    time.sleep(t/counts)          

                    
                p0=pf
                presscontype='Stable'
            if pressswitch():
                console_out("Stop gas control")
                break
                    
                                    
                
    #def mass_check_thread(press_control,p0,pf,mass_Flow):
     #   if press_control="Stable"#self.acquisition_instance = threading.Thread(target=self.acquisition_thread, daemon=True)
      #      current_thread = threading.currentThread()

        #while(getattr(current_thread, "active", True)):
            
        #p0 = _GUI.P.get()
        #p00 = _IO.pressure.get()
        #mass_Step = 0.05
        
        
